from openpyxl import load_workbook


def get_row_count(file, sheet_name):

    workbook = load_workbook(file)

    sheet = workbook[sheet_name]

    return sheet.max_row


def read_data(file, sheet_name, row, column):

    workbook = load_workbook(file)

    sheet = workbook[sheet_name]

    return sheet.cell(row=row, column=column).value